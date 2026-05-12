from io import StringIO
from pathlib import Path

import openpyxl
import pytest
from rich.console import Console
from typer.testing import CliRunner

import safe.cli.board as board_module
import safe.cli.state as state
from safe.cli.main import app
from safe.store.db import get_db
from safe.store.repos import get_repos

runner = CliRunner()


@pytest.fixture(autouse=True)
def reset_state():
    state.db_path = None
    yield
    state.db_path = None


@pytest.fixture
def db_path(tmp_path: Path) -> Path:
    return tmp_path / "test.json"


@pytest.fixture(autouse=True)
def patch_console(monkeypatch):
    buf = StringIO()
    test_console = Console(file=buf, highlight=False, markup=False, width=200)
    monkeypatch.setattr(board_module, "console", test_console)
    yield buf


def invoke(db_path: Path, *args):
    return runner.invoke(app, ["--db-path", str(db_path)] + list(args))


def repos_for(db_path: Path):
    return get_repos(get_db(db_path))


def _setup(db_path):
    """Create ART, PI, 2 teams, 2 iterations, 2 features (one with stories, one without)."""
    invoke(db_path, "art", "create", "--name", "ART")
    art_id = repos_for(db_path).arts.get_all()[0].id

    invoke(
        db_path,
        "pi",
        "create",
        "--name",
        "PI 1",
        "--art-id",
        art_id,
        "--start",
        "2026-01-05",
        "--end",
        "2026-03-27",
    )
    pi_id = repos_for(db_path).pis.get_all()[0].id

    invoke(db_path, "team", "create", "--name", "Alpha", "--members", "6")
    invoke(db_path, "team", "create", "--name", "Beta", "--members", "5")
    teams = repos_for(db_path).teams.get_all()
    alpha_id = next(t.id for t in teams if t.name == "Alpha")
    beta_id = next(t.id for t in teams if t.name == "Beta")

    invoke(
        db_path,
        "pi",
        "iteration",
        "add",
        "--pi-id",
        pi_id,
        "--number",
        "1",
        "--start",
        "2026-01-05",
        "--end",
        "2026-01-16",
    )
    invoke(
        db_path,
        "pi",
        "iteration",
        "add",
        "--pi-id",
        pi_id,
        "--number",
        "2",
        "--start",
        "2026-01-19",
        "--end",
        "2026-01-30",
    )
    iterations = repos_for(db_path).iterations.find(pi_id=pi_id)
    iter1_id = next(i.id for i in iterations if i.number == 1)

    # Feature with stories in iter1 → should appear under I1 for Alpha
    invoke(
        db_path,
        "feature",
        "add",
        "--name",
        "Auth Service",
        "--pi-id",
        pi_id,
        "--user-value",
        "8",
        "--time-crit",
        "5",
        "--risk-reduction",
        "3",
        "--job-size",
        "4",
    )
    feat1_id = repos_for(db_path).features.get_all()[0].id
    invoke(db_path, "feature", "assign", feat1_id, "--team-id", alpha_id)
    invoke(
        db_path,
        "story",
        "add",
        "--name",
        "Login flow",
        "--feature-id",
        feat1_id,
        "--team-id",
        alpha_id,
        "--points",
        "3",
        "--iteration-id",
        iter1_id,
    )

    # Feature without stories → should appear in Unplanned for Beta
    invoke(
        db_path,
        "feature",
        "add",
        "--name",
        "Reporting Dashboard",
        "--pi-id",
        pi_id,
        "--user-value",
        "6",
        "--time-crit",
        "4",
        "--risk-reduction",
        "2",
        "--job-size",
        "5",
    )
    feat2_id = next(
        f.id for f in repos_for(db_path).features.get_all() if f.name == "Reporting Dashboard"
    )
    invoke(db_path, "feature", "assign", feat2_id, "--team-id", beta_id)

    return pi_id, alpha_id, beta_id, iter1_id, feat1_id, feat2_id


class TestBoardShow:
    def test_exit_code_success(self, db_path, patch_console):
        pi_id, *_ = _setup(db_path)
        result = invoke(db_path, "board", "show", "--pi-id", pi_id)
        assert result.exit_code == 0

    def test_shows_team_names(self, db_path, patch_console):
        pi_id, *_ = _setup(db_path)
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "board", "show", "--pi-id", pi_id)
        output = patch_console.getvalue()
        assert "Alpha" in output
        assert "Beta" in output

    def test_shows_feature_in_iteration_column(self, db_path, patch_console):
        pi_id, *_ = _setup(db_path)
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "board", "show", "--pi-id", pi_id)
        output = patch_console.getvalue()
        assert "Auth Service" in output
        assert "I1" in output

    def test_unplanned_column_present(self, db_path, patch_console):
        pi_id, *_ = _setup(db_path)
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "board", "show", "--pi-id", pi_id)
        assert "Unplanned" in patch_console.getvalue()

    def test_unassigned_feature_in_unplanned(self, db_path, patch_console):
        pi_id, *_ = _setup(db_path)
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "board", "show", "--pi-id", pi_id)
        assert "Reporting Dashboard" in patch_console.getvalue()

    def test_shows_dependencies_section(self, db_path, patch_console):
        pi_id, _alpha_id, _beta_id, _iter1_id, feat1_id, feat2_id = _setup(db_path)
        invoke(
            db_path,
            "dependency",
            "add",
            "--description",
            "Shared API contract",
            "--pi-id",
            pi_id,
            "--from-feature-id",
            feat1_id,
            "--to-feature-id",
            feat2_id,
        )
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "board", "show", "--pi-id", pi_id)
        output = patch_console.getvalue()
        assert "Dependencies" in output
        assert "Shared API contract" in output

    def test_no_dependencies_section_when_empty(self, db_path, patch_console):
        pi_id, *_ = _setup(db_path)
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "board", "show", "--pi-id", pi_id)
        assert "Dependencies" not in patch_console.getvalue()

    def test_unknown_pi_exits_1(self, db_path, patch_console):
        result = invoke(db_path, "board", "show", "--pi-id", "bad-id")
        assert result.exit_code == 1

    def test_empty_pi_shows_message(self, db_path, patch_console):
        invoke(db_path, "art", "create", "--name", "ART")
        art_id = repos_for(db_path).arts.get_all()[0].id
        invoke(
            db_path,
            "pi",
            "create",
            "--name",
            "Empty PI",
            "--art-id",
            art_id,
            "--start",
            "2026-01-05",
            "--end",
            "2026-03-27",
        )
        pi_id = repos_for(db_path).pis.get_all()[0].id
        patch_console.truncate(0)
        patch_console.seek(0)
        invoke(db_path, "board", "show", "--pi-id", pi_id)
        assert "No features" in patch_console.getvalue()


class TestBoardExport:
    def test_creates_file(self, db_path, patch_console, tmp_path):
        pi_id, *_ = _setup(db_path)
        output = tmp_path / "board.xlsx"
        result = invoke(db_path, "board", "export", "--pi-id", pi_id, "--output", str(output))
        assert result.exit_code == 0
        assert output.exists()

    def test_workbook_contains_team_names(self, db_path, patch_console, tmp_path):
        pi_id, *_ = _setup(db_path)
        output = tmp_path / "board.xlsx"
        invoke(db_path, "board", "export", "--pi-id", pi_id, "--output", str(output))
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
        assert any("Alpha" in v for v in values)
        assert any("Beta" in v for v in values)

    def test_workbook_contains_feature_names(self, db_path, patch_console, tmp_path):
        pi_id, *_ = _setup(db_path)
        output = tmp_path / "board.xlsx"
        invoke(db_path, "board", "export", "--pi-id", pi_id, "--output", str(output))
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
        assert any("Auth Service" in v for v in values)
        assert any("Reporting Dashboard" in v for v in values)

    def test_dependencies_sheet_created(self, db_path, patch_console, tmp_path):
        pi_id, _alpha_id, _beta_id, _iter1_id, feat1_id, feat2_id = _setup(db_path)
        invoke(
            db_path,
            "dependency",
            "add",
            "--description",
            "Platform contract",
            "--pi-id",
            pi_id,
            "--from-feature-id",
            feat1_id,
            "--to-feature-id",
            feat2_id,
        )
        output = tmp_path / "board.xlsx"
        invoke(db_path, "board", "export", "--pi-id", pi_id, "--output", str(output))
        wb = openpyxl.load_workbook(output)
        assert "Dependencies" in wb.sheetnames

    def test_unknown_pi_exits_1(self, db_path, patch_console, tmp_path):
        output = tmp_path / "board.xlsx"
        result = invoke(db_path, "board", "export", "--pi-id", "bad-id", "--output", str(output))
        assert result.exit_code == 1


class TestBoardLogic:
    """Unit tests for the board logic without CLI overhead."""

    def test_feature_without_team_excluded_from_grid(self, db_path, patch_console):
        from safe.logic.board import build_board
        from safe.models.backlog import Feature

        f = Feature(
            id="f1",
            name="F1",
            team_id=None,
            pi_id="pi1",
            user_business_value=5,
            time_criticality=5,
            risk_reduction_opportunity_enablement=5,
            job_size=5,
        )
        grid = build_board([f])
        assert grid == {}

    def test_feature_placed_in_correct_iteration(self, db_path, patch_console):
        from safe.logic.board import build_board
        from safe.models.backlog import Feature

        f = Feature(
            id="f1",
            name="F1",
            team_id="t1",
            iteration_id="i-override",
            user_business_value=5,
            time_criticality=5,
            risk_reduction_opportunity_enablement=5,
            job_size=5,
        )
        grid = build_board([f])
        assert "i-override" in grid["t1"]

    def test_iteration_id_none_places_feature_in_unplanned(self, db_path, patch_console):
        from safe.logic.board import build_board
        from safe.models.backlog import Feature

        f = Feature(
            id="f1",
            name="F1",
            team_id="t1",
            iteration_id=None,
            user_business_value=5,
            time_criticality=5,
            risk_reduction_opportunity_enablement=5,
            job_size=5,
        )
        grid = build_board([f])
        assert None in grid["t1"]
        assert grid["t1"][None] == [f]
